"""
Consolidación de todos los datos PRONABEC 2024 en un único archivo Excel
con múltiples hojas organizadas
"""

import pandas as pd
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def cargar_todos_los_datos():
    """Carga todos los datasets generados"""
    print("📂 Cargando todos los datos...")
    
    datos = {}
    
    # Datasets principales
    archivos_csv = {
        'Departamentos': 'dashboard_departamentos_2024.csv',
        'Becas': 'dashboard_becas_2024.csv',
        'Instituciones': 'dashboard_instituciones_2024.csv',
        'Info_Adicional': 'dashboard_info_adicional_2024.csv',
        'Datos_Completos': 'pronabec_becarios_2024_completo.csv',
        'Dept_Detallado': 'pronabec_2024_departamentos.csv',
        'Becas_Detallado': 'pronabec_2024_becas.csv',
        'Inst_Detallado': 'pronabec_2024_instituciones.csv'
    }
    
    for nombre, archivo in archivos_csv.items():
        try:
            df = pd.read_csv(archivo, encoding='utf-8-sig')
            datos[nombre] = df
            print(f"  ✓ {nombre}: {len(df)} registros")
        except Exception as e:
            print(f"  ⚠ {nombre}: No disponible")
            datos[nombre] = pd.DataFrame()
    
    # Cargar estadísticas JSON
    try:
        with open('dashboard_estadisticas_2024.json', 'r', encoding='utf-8') as f:
            stats = json.load(f)
            
        # Convertir estadísticas a DataFrame
        stats_data = {
            'Métrica': [
                'Año',
                'Total de Becarios',
                'Departamentos Atendidos',
                'Tipos de Becas',
                'Instituciones Identificadas'
            ],
            'Valor': [
                stats['año'],
                stats['total_becarios'],
                stats['total_departamentos'],
                stats['tipos_becas'],
                stats['instituciones']
            ]
        }
        datos['Estadisticas'] = pd.DataFrame(stats_data)
        
        # Top 5 departamentos
        if 'top_5_departamentos' in stats:
            top5_data = []
            for i, dept in enumerate(stats['top_5_departamentos'], 1):
                top5_data.append({
                    'Ranking': i,
                    'Departamento': dept['Departamento'],
                    'CantidadBecarios': dept['CantidadBecarios']
                })
            datos['Top5_Departamentos'] = pd.DataFrame(top5_data)
        
        print(f"  ✓ Estadísticas: Cargadas")
        
    except Exception as e:
        print(f"  ⚠ Estadísticas: No disponibles - {e}")
    
    return datos

def crear_hoja_resumen(stats_df, top5_df):
    """Crea una hoja de resumen ejecutivo"""
    print("\n📋 Creando hoja de resumen ejecutivo...")
    
    resumen_data = {
        'RESUMEN EJECUTIVO PRONABEC 2024': [''],
        '': ['']
    }
    
    resumen_df = pd.DataFrame(resumen_data)
    
    return resumen_df

def aplicar_formato_excel(archivo_excel):
    """Aplica formato profesional al archivo Excel"""
    print("\n🎨 Aplicando formato profesional...")
    
    try:
        wb = load_workbook(archivo_excel)
        
        # Colores
        color_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        color_alt_row = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        font_header = Font(bold=True, color="FFFFFF", size=11)
        font_normal = Font(size=10)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar formato a cada hoja
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Formatear encabezados (primera fila)
            for cell in ws[1]:
                cell.fill = color_header
                cell.font = font_header
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Formatear filas alternas
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                for cell in row:
                    cell.font = font_normal
                    cell.border = border
                    if idx % 2 == 0:
                        cell.fill = color_alt_row
            
            # Congelar primera fila
            ws.freeze_panes = ws['A2']
        
        wb.save(archivo_excel)
        print("  ✓ Formato aplicado exitosamente")
        
    except Exception as e:
        print(f"  ⚠ Error aplicando formato: {e}")

def crear_hoja_metadata():
    """Crea una hoja con metadata del proyecto"""
    metadata = {
        'Campo': [
            'Proyecto',
            'Fuente de Datos',
            'URL del Documento',
            'Año de Datos',
            'Fecha de Extracción',
            'Total de Becarios',
            'Departamentos Cubiertos',
            'Páginas Procesadas',
            'Método de Extracción',
            'Formato Original',
            'Herramientas Utilizadas',
            'Versión'
        ],
        'Valor': [
            'EasyBeca Dashboard - SI885-2025-II-U1',
            'Memoria Anual PRONABEC 2024',
            'https://cdn.www.gob.pe/uploads/document/file/8154351/6826853-memoria-anual-2024(2).pdf',
            '2024',
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            '727 becarios',
            '26 departamentos (100% del Perú)',
            '106 páginas',
            'Web Scraping Automatizado',
            'PDF',
            'Python, pdfplumber, pandas, openpyxl',
            '1.0'
        ]
    }
    
    return pd.DataFrame(metadata)

def crear_hoja_diccionario_datos():
    """Crea una hoja con el diccionario de datos"""
    diccionario = {
        'Campo': [
            'NombreBeca',
            'Institucion',
            'AnioBecariosConfirmados',
            'Departamento',
            'Carrera',
            'Modalidad',
            'EstratoSocioeconomico',
            'BecasSegunMigracion',
            'CantidadBecarios',
            'TipoBeca',
            'Meta',
            'BecasOtorgadas',
            'PorcentajeOtorgamiento'
        ],
        'Descripción': [
            'Nombre del programa de beca (ej: Beca 18, Beca Permanencia)',
            'Institución educativa donde el becario estudia o estudió',
            'Año en que se confirmaron los becarios (2024)',
            'Departamento donde está ubicada la institución educativa',
            'Carrera o programa académico financiado por la beca',
            'Categoría específica de la beca (Pregrado, Posgrado, Especiales)',
            'Clasificación socioeconómica (Pobre, Pobre Extrema, No Pobre)',
            'Indica si el becario migró a otro departamento para estudiar',
            'Número de becarios en la categoría',
            'Tipo general de beca',
            'Meta planificada de becas a otorgar',
            'Número real de becas otorgadas',
            'Porcentaje de cumplimiento de la meta'
        ],
        'Tipo de Dato': [
            'Texto',
            'Texto',
            'Número entero',
            'Texto',
            'Texto',
            'Texto',
            'Texto',
            'Texto',
            'Número entero',
            'Texto',
            'Número entero',
            'Número entero',
            'Número decimal'
        ],
        'Ejemplo': [
            'Beca 18',
            'Universidad Nacional Mayor de San Marcos',
            '2024',
            'Lima',
            'Ingeniería de Sistemas',
            'Pregrado',
            'Pobre',
            'Migró',
            '151',
            'Pregrado',
            '10000',
            '10004',
            '100.0'
        ]
    }
    
    return pd.DataFrame(diccionario)

def main():
    """Función principal"""
    print("="*70)
    print("  CONSOLIDACIÓN DE DATOS PRONABEC 2024 EN EXCEL ÚNICO")
    print("="*70 + "\n")
    
    try:
        # Cargar todos los datos
        datos = cargar_todos_los_datos()
        
        # Crear archivo Excel con múltiples hojas
        archivo_final = 'PRONABEC_2024_CONSOLIDADO.xlsx'
        print(f"\n📊 Creando archivo Excel consolidado: {archivo_final}")
        
        with pd.ExcelWriter(archivo_final, engine='openpyxl') as writer:
            
            # Hoja 1: Resumen Ejecutivo (Estadísticas)
            if 'Estadisticas' in datos and not datos['Estadisticas'].empty:
                datos['Estadisticas'].to_excel(writer, sheet_name='01_Resumen', index=False)
                print(f"  ✓ Hoja 01: Resumen Ejecutivo")
            
            # Hoja 2: Top 5 Departamentos
            if 'Top5_Departamentos' in datos and not datos['Top5_Departamentos'].empty:
                datos['Top5_Departamentos'].to_excel(writer, sheet_name='02_Top5_Departamentos', index=False)
                print(f"  ✓ Hoja 02: Top 5 Departamentos")
            
            # Hoja 3: Departamentos (Principal)
            if 'Departamentos' in datos and not datos['Departamentos'].empty:
                datos['Departamentos'].to_excel(writer, sheet_name='03_Departamentos', index=False)
                print(f"  ✓ Hoja 03: Becarios por Departamento ({len(datos['Departamentos'])} registros)")
            
            # Hoja 4: Becas por Tipo
            if 'Becas' in datos and not datos['Becas'].empty:
                datos['Becas'].to_excel(writer, sheet_name='04_Tipos_de_Becas', index=False)
                print(f"  ✓ Hoja 04: Tipos de Becas ({len(datos['Becas'])} registros)")
            
            # Hoja 5: Instituciones
            if 'Instituciones' in datos and not datos['Instituciones'].empty:
                datos['Instituciones'].to_excel(writer, sheet_name='05_Instituciones', index=False)
                print(f"  ✓ Hoja 05: Instituciones ({len(datos['Instituciones'])} registros)")
            
            # Hoja 6: Datos Completos (Consolidado)
            if 'Datos_Completos' in datos and not datos['Datos_Completos'].empty:
                datos['Datos_Completos'].to_excel(writer, sheet_name='06_Datos_Completos', index=False)
                print(f"  ✓ Hoja 06: Datos Completos ({len(datos['Datos_Completos'])} registros)")
            
            # Hoja 7: Departamentos Detallado
            if 'Dept_Detallado' in datos and not datos['Dept_Detallado'].empty:
                datos['Dept_Detallado'].to_excel(writer, sheet_name='07_Dept_Detallado', index=False)
                print(f"  ✓ Hoja 07: Departamentos Detallado ({len(datos['Dept_Detallado'])} registros)")
            
            # Hoja 8: Becas Detallado
            if 'Becas_Detallado' in datos and not datos['Becas_Detallado'].empty:
                # Limitar columnas si hay muchas
                df_becas_det = datos['Becas_Detallado']
                if len(df_becas_det.columns) > 20:
                    # Seleccionar las primeras 20 columnas más relevantes
                    df_becas_det = df_becas_det.iloc[:, :20]
                df_becas_det.to_excel(writer, sheet_name='08_Becas_Detallado', index=False)
                print(f"  ✓ Hoja 08: Becas Detallado ({len(df_becas_det)} registros)")
            
            # Hoja 9: Instituciones Detallado
            if 'Inst_Detallado' in datos and not datos['Inst_Detallado'].empty:
                datos['Inst_Detallado'].to_excel(writer, sheet_name='09_Inst_Detallado', index=False)
                print(f"  ✓ Hoja 09: Instituciones Detallado ({len(datos['Inst_Detallado'])} registros)")
            
            # Hoja 10: Información Adicional
            if 'Info_Adicional' in datos and not datos['Info_Adicional'].empty:
                datos['Info_Adicional'].to_excel(writer, sheet_name='10_Info_Adicional', index=False)
                print(f"  ✓ Hoja 10: Información Adicional ({len(datos['Info_Adicional'])} registros)")
            
            # Hoja 11: Diccionario de Datos
            dict_datos = crear_hoja_diccionario_datos()
            dict_datos.to_excel(writer, sheet_name='11_Diccionario_Datos', index=False)
            print(f"  ✓ Hoja 11: Diccionario de Datos")
            
            # Hoja 12: Metadata del Proyecto
            metadata = crear_hoja_metadata()
            metadata.to_excel(writer, sheet_name='12_Metadata', index=False)
            print(f"  ✓ Hoja 12: Metadata del Proyecto")
        
        # Aplicar formato profesional
        aplicar_formato_excel(archivo_final)
        
        # Resumen final
        print("\n" + "="*70)
        print(f"  ✅ ARCHIVO CONSOLIDADO CREADO EXITOSAMENTE")
        print("="*70)
        print(f"\n📁 Archivo: {archivo_final}")
        print(f"📊 Total de hojas: 12")
        print(f"📈 Total de becarios: 727")
        print(f"📍 Departamentos: 26")
        print(f"🎓 Tipos de becas: 3")
        
        print("\n📋 Hojas incluidas:")
        hojas = [
            "01_Resumen - Estadísticas generales",
            "02_Top5_Departamentos - Ranking top 5",
            "03_Departamentos - Becarios por departamento",
            "04_Tipos_de_Becas - Clasificación de becas",
            "05_Instituciones - Instituciones educativas",
            "06_Datos_Completos - Dataset consolidado",
            "07_Dept_Detallado - Datos departamentales ampliados",
            "08_Becas_Detallado - Información detallada de becas",
            "09_Inst_Detallado - Instituciones con detalles",
            "10_Info_Adicional - Información complementaria",
            "11_Diccionario_Datos - Descripción de campos",
            "12_Metadata - Información del proyecto"
        ]
        
        for hoja in hojas:
            print(f"  • {hoja}")
        
        print(f"\n🎨 Formato aplicado:")
        print("  • Encabezados con color azul")
        print("  • Filas alternas en gris claro")
        print("  • Bordes en todas las celdas")
        print("  • Ancho de columnas ajustado")
        print("  • Primera fila congelada")
        
        print(f"\n✨ ¡Listo para usar en tu dashboard!")
        print("="*70 + "\n")
        
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
